"""知识治理服务验收测试（服务层 + 临时 SQLite，不依赖 Milvus/docker）。

覆盖：sync 补建、列表/详情合并元数据、路径不泄露、PATCH 更新与非法值、
伪造 db_id/file_id、下载权限矩阵、路径穿越/软链接逃逸/URL 拒绝、
usage_count 服务端自增、JSON/XLSX 导出不含 path。
"""

import hashlib
import os
import tempfile
import threading
import time
import unittest
from contextlib import contextmanager
from datetime import datetime
from pathlib import Path
from unittest import mock

from sqlalchemy import create_engine, text
from sqlalchemy.exc import SQLAlchemyError as _SQAError
from sqlalchemy.orm import sessionmaker

import server.models.governance_model  # noqa: F401
import server.models.kb_models  # noqa: F401
from server.models import Base
from server.models.governance_model import KnowledgeDocumentVersion, KnowledgeGovernance
from server.models.kb_models import KnowledgeDatabase, KnowledgeFile
from server.services import governance_service
from server.services.governance_service import (
    GovernanceError,
    GovernanceForbidden,
    GovernanceNotFound,
)

_ORIG_DATA_ROOT = governance_service.DATA_ROOT
_ORIG_PROJECT_ROOT = governance_service._PROJECT_ROOT


@contextmanager
def _temp_db():
    with tempfile.TemporaryDirectory() as tmp:
        db_path = Path(tmp) / "test.db"
        engine = create_engine(
            f"sqlite:///{db_path}",
            connect_args={"check_same_thread": False},
        )
        # 注意：不开启 foreign_keys=ON。
        # knowledge_nodes.file_id 引用非唯一的 knowledge_files.file_id，
        # 属于既有模型设计，SQLite 开启外键后会报 mismatch。
        Base.metadata.create_all(engine)
        Session = sessionmaker(bind=engine)
        session = Session()
        try:
            yield engine, session
        finally:
            session.close()
            engine.dispose()


def _set_roots(data_root):
    governance_service.DATA_ROOT = data_root
    governance_service._PROJECT_ROOT = data_root


def _restore_roots():
    governance_service.DATA_ROOT = _ORIG_DATA_ROOT
    governance_service._PROJECT_ROOT = _ORIG_PROJECT_ROOT


def _seed(session, data_root, db_id="kb_test", names=("a.pdf", "b.txt"), status="done"):
    db = KnowledgeDatabase(db_id=db_id, name="测试知识库")
    session.add(db)
    session.flush()
    for name in names:
        p = os.path.join(data_root, name)
        Path(p).write_bytes(("content " + name).encode("utf-8"))
        ext = os.path.splitext(name)[1].lstrip(".")
        session.add(
            KnowledgeFile(
                database_id=db_id,
                file_id="file_" + name.split(".")[0],
                filename=name,
                path=p,
                file_type=ext,
                status=status,
            )
        )
    session.commit()


def _blob_files(data_root):
    """收集内容寻址 blob 下的已发布文件（排除请求级暂存目录）。"""
    blobs = os.path.join(data_root, "knowledge_versions", "_blobs")
    if not os.path.isdir(blobs):
        return []
    out = []
    for root, _dirs, files in os.walk(blobs):
        if "_staging" in root:
            continue
        for name in files:
            out.append(os.path.join(root, name))
    return out


def _seed_file(session, data_root, db_id, file_id, filename, path, file_type="pdf"):
    db = session.query(KnowledgeDatabase).filter_by(db_id=db_id).first()
    if db is None:
        db = KnowledgeDatabase(db_id=db_id, name=db_id)
        session.add(db)
        session.flush()
    session.add(
        KnowledgeFile(
            database_id=db_id,
            file_id=file_id,
            filename=filename,
            path=path,
            file_type=file_type,
            status="done",
        )
    )
    session.commit()


class _User:
    def __init__(self, role):
        self.id = 1
        self.role = role


class GovernanceServiceTests(unittest.TestCase):
    def test_sync_creates_default_governance_rows(self):
        with _temp_db() as (engine, session):
            with tempfile.TemporaryDirectory() as data_root:
                _seed(session, data_root)
                result = governance_service.sync_governance(session, "kb_test")
                self.assertEqual(result["total"], 2)
                self.assertEqual(result["created"], 2)
                self.assertEqual(result["updated"], 0)
                rows = (
                    session.query(KnowledgeGovernance)
                    .filter_by(db_id="kb_test")
                    .all()
                )
                self.assertEqual(len(rows), 2)
                for r in rows:
                    self.assertEqual(r.confidentiality, "internal")
                    self.assertEqual(r.download_allowed, 1)
                    self.assertIsNotNone(r.source_updated_at)

    def test_sync_idempotent(self):
        with _temp_db() as (engine, session):
            with tempfile.TemporaryDirectory() as data_root:
                _seed(session, data_root)
                governance_service.sync_governance(session, "kb_test")
                result = governance_service.sync_governance(session, "kb_test")
                self.assertEqual(result["created"], 0)

    def test_sync_missing_database_raises_not_found(self):
        with _temp_db() as (engine, session):
            with self.assertRaises(GovernanceNotFound):
                governance_service.sync_governance(session, "kb_missing")

    def test_list_documents_merges_metadata_without_path(self):
        with _temp_db() as (engine, session):
            with tempfile.TemporaryDirectory() as data_root:
                _seed(session, data_root)
                governance_service.sync_governance(session, "kb_test")
                data = governance_service.list_documents(session, "kb_test")
                self.assertEqual(data["total"], 2)
                self.assertEqual(len(data["items"]), 2)
                for it in data["items"]:
                    self.assertNotIn("path", it, "列表响应不得泄露文件路径")
                    self.assertIn("filename", it)
                    self.assertIn("confidentiality", it)
                    self.assertIn("download_allowed", it)
                    self.assertIn("usage_count", it)
                    self.assertIn("node_count", it)

    def test_list_filters_and_pagination(self):
        with _temp_db() as (engine, session):
            with tempfile.TemporaryDirectory() as data_root:
                _seed(session, data_root)
                governance_service.sync_governance(session, "kb_test")
                gov = (
                    session.query(KnowledgeGovernance)
                    .filter_by(db_id="kb_test", file_id="file_a")
                    .first()
                )
                gov.knowledge_type = "报告"
                gov.confidentiality = "restricted"
                session.commit()
                # 按类型筛选
                data = governance_service.list_documents(
                    session, "kb_test", knowledge_type="报告"
                )
                self.assertEqual(data["total"], 1)
                self.assertEqual(data["items"][0]["file_id"], "file_a")
                # 按关键字筛选
                data = governance_service.list_documents(session, "kb_test", keyword="b.txt")
                self.assertEqual(data["total"], 1)
                self.assertEqual(data["items"][0]["filename"], "b.txt")
                # 分页
                data = governance_service.list_documents(
                    session, "kb_test", page=1, page_size=1
                )
                self.assertEqual(len(data["items"]), 1)
                self.assertEqual(data["page"], 1)

    def test_get_document_increments_usage(self):
        with _temp_db() as (engine, session):
            with tempfile.TemporaryDirectory() as data_root:
                _seed(session, data_root)
                doc = governance_service.get_document(session, "kb_test", "file_a")
                self.assertEqual(doc["filename"], "a.pdf")
                self.assertNotIn("path", doc)
                self.assertEqual(doc["usage_count"], 1)
                doc2 = governance_service.get_document(session, "kb_test", "file_a")
                self.assertEqual(doc2["usage_count"], 2)

    def test_patch_updates_governance_fields(self):
        with _temp_db() as (engine, session):
            with tempfile.TemporaryDirectory() as data_root:
                _seed(session, data_root)
                doc = governance_service.update_governance(
                    session,
                    "kb_test",
                    "file_a",
                    {
                        "domain": "石油储运",
                        "knowledge_type": "标准",
                        "confidentiality": "restricted",
                        "tags": ["规范", "安全"],
                        "download_allowed": False,
                        "owner_department": "工程部",
                    },
                )
                self.assertEqual(doc["domain"], "石油储运")
                self.assertEqual(doc["knowledge_type"], "标准")
                self.assertEqual(doc["confidentiality"], "restricted")
                self.assertEqual(doc["tags"], ["规范", "安全"])
                self.assertIs(doc["download_allowed"], False)
                self.assertEqual(doc["owner_department"], "工程部")
                # 持久化
                got = governance_service.get_document(session, "kb_test", "file_a")
                self.assertEqual(got["confidentiality"], "restricted")

    def test_patch_null_clears_optional_fields(self):
        """P2-2 验收：显式 null 清空可选字段，重新加载后旧值不保留。"""
        with _temp_db() as (engine, session):
            with tempfile.TemporaryDirectory() as data_root:
                _seed(session, data_root)
                governance_service.update_governance(
                    session,
                    "kb_test",
                    "file_a",
                    {
                        "domain": "石油储运",
                        "knowledge_type": "标准",
                        "tags": ["规范", "安全"],
                        "owner_department": "工程部",
                        "source_updated_at": datetime(2024, 1, 1, 12, 0, 0),
                    },
                )
                # 明确清空（等价于 PATCH 发送 null）
                governance_service.update_governance(
                    session,
                    "kb_test",
                    "file_a",
                    {"domain": None, "tags": None, "owner_department": None, "knowledge_type": None},
                )
                reloaded = governance_service.get_document(session, "kb_test", "file_a")
                self.assertIsNone(reloaded["domain"])
                self.assertIsNone(reloaded["knowledge_type"])
                self.assertEqual(reloaded["tags"], [])
                self.assertIsNone(reloaded["owner_department"])
                # 未提交的字段不受影响
                self.assertEqual(reloaded["confidentiality"], "internal")

    def test_patch_rejects_invalid_values(self):
        with _temp_db() as (engine, session):
            with tempfile.TemporaryDirectory() as data_root:
                _seed(session, data_root)
                with self.assertRaises(GovernanceError):
                    governance_service.update_governance(
                        session, "kb_test", "file_a", {"confidentiality": "top_secret"}
                    )
                with self.assertRaises(GovernanceError):
                    governance_service.update_governance(
                        session, "kb_test", "file_a", {"knowledge_type": "其他类型"}
                    )
                with self.assertRaises(GovernanceError):
                    governance_service.update_governance(
                        session, "kb_test", "file_a", {"tags": "not-a-list"}
                    )
                with self.assertRaises(GovernanceError):
                    governance_service.update_governance(
                        session, "kb_test", "file_a", {"mystery_field": 1}
                    )

    def test_forged_ids_rejected(self):
        with _temp_db() as (engine, session):
            with tempfile.TemporaryDirectory() as data_root:
                _seed(session, data_root)
                with self.assertRaises(GovernanceNotFound):
                    governance_service.get_document(session, "kb_missing", "file_a")
                with self.assertRaises(GovernanceNotFound):
                    governance_service.get_document(session, "kb_test", "file_missing")
                with self.assertRaises(GovernanceError):
                    governance_service.get_document(session, "kb_test", "../etc/passwd")
                with self.assertRaises(GovernanceError):
                    governance_service.get_document(session, "kb_test\\evil", "file_a")

    def test_download_permission_matrix(self):
        with _temp_db() as (engine, session):
            with tempfile.TemporaryDirectory() as data_root:
                _set_roots(data_root)
                try:
                    _seed(session, data_root)
                    # 默认 internal + download_allowed=1 → 普通用户可下载
                    info = governance_service.resolve_download(
                        session, "kb_test", "file_a", _User("user")
                    )
                    self.assertEqual(info["filename"], "a.pdf")
                    self.assertGreater(info["size_bytes"], 0)
                    self.assertTrue(os.path.isfile(info["abs_path"]))
                    # restricted → 非 superadmin 拒绝
                    governance_service.update_governance(
                        session, "kb_test", "file_a", {"confidentiality": "restricted"}
                    )
                    with self.assertRaises(GovernanceForbidden):
                        governance_service.resolve_download(
                            session, "kb_test", "file_a", _User("admin")
                        )
                    info = governance_service.resolve_download(
                        session, "kb_test", "file_a", _User("superadmin")
                    )
                    self.assertEqual(info["filename"], "a.pdf")
                    # download_allowed=0 → 所有人（含 superadmin）拒绝
                    governance_service.update_governance(
                        session, "kb_test", "file_b", {"download_allowed": False}
                    )
                    with self.assertRaises(GovernanceForbidden):
                        governance_service.resolve_download(
                            session, "kb_test", "file_b", _User("user")
                        )
                    with self.assertRaises(GovernanceForbidden):
                        governance_service.resolve_download(
                            session, "kb_test", "file_b", _User("superadmin")
                        )
                finally:
                    _restore_roots()

    def test_download_increments_usage_count(self):
        with _temp_db() as (engine, session):
            with tempfile.TemporaryDirectory() as data_root:
                _set_roots(data_root)
                try:
                    _seed(session, data_root)
                    governance_service.resolve_download(
                        session, "kb_test", "file_a", _User("user")
                    )
                    governance_service.resolve_download(
                        session, "kb_test", "file_a", _User("user")
                    )
                    got = governance_service.get_document(
                        session, "kb_test", "file_a", increment_usage=False
                    )
                    self.assertEqual(got["usage_count"], 2)
                finally:
                    _restore_roots()

    def test_path_traversal_rejected(self):
        with _temp_db() as (engine, session):
            with tempfile.TemporaryDirectory() as data_root:
                _set_roots(data_root)
                try:
                    evil = os.path.join(data_root, "..", "outside.txt")
                    Path(evil).write_bytes(b"secret")
                    _seed_file(
                        session, data_root, "kb_test", "file_evil", "evil.txt", evil
                    )
                    with self.assertRaises(GovernanceNotFound):
                        governance_service.resolve_download(
                            session, "kb_test", "file_evil", _User("user")
                        )
                finally:
                    _restore_roots()

    def test_absolute_path_outside_root_rejected(self):
        with _temp_db() as (engine, session):
            with tempfile.TemporaryDirectory() as data_root:
                with tempfile.TemporaryDirectory() as outside:
                    _set_roots(data_root)
                    try:
                        evil = os.path.join(outside, "secret.txt")
                        Path(evil).write_bytes(b"secret")
                        _seed_file(
                            session, data_root, "kb_test", "file_abs", "secret.txt", evil
                        )
                        with self.assertRaises(GovernanceNotFound):
                            governance_service.resolve_download(
                                session, "kb_test", "file_abs", _User("user")
                            )
                    finally:
                        _restore_roots()

    def test_url_path_rejected(self):
        with _temp_db() as (engine, session):
            with tempfile.TemporaryDirectory() as data_root:
                _set_roots(data_root)
                try:
                    _seed_file(
                        session,
                        data_root,
                        "kb_test",
                        "file_url",
                        "remote.pdf",
                        "http://10.16.33.2:8002/upload/remote.pdf",
                    )
                    with self.assertRaises(GovernanceNotFound):
                        governance_service.resolve_download(
                            session, "kb_test", "file_url", _User("user")
                        )
                finally:
                    _restore_roots()

    def test_symlink_escape_rejected(self):
        with _temp_db() as (engine, session):
            with tempfile.TemporaryDirectory() as data_root:
                with tempfile.TemporaryDirectory() as outside:
                    _set_roots(data_root)
                    try:
                        target = os.path.join(outside, "real.txt")
                        Path(target).write_bytes(b"secret")
                        link = os.path.join(data_root, "link.txt")
                        try:
                            os.symlink(target, link)
                        except (OSError, NotImplementedError):
                            self.skipTest("当前环境不允许创建软链接")
                        _seed_file(
                            session, data_root, "kb_test", "file_link", "link.txt", link
                        )
                        with self.assertRaises(GovernanceNotFound):
                            governance_service.resolve_download(
                                session, "kb_test", "file_link", _User("user")
                            )
                    finally:
                        _restore_roots()

    def test_export_json_no_path_no_secrets(self):
        with _temp_db() as (engine, session):
            with tempfile.TemporaryDirectory() as data_root:
                _seed(session, data_root)
                governance_service.update_governance(
                    session,
                    "kb_test",
                    "file_a",
                    {"domain": "安全", "confidentiality": "restricted"},
                )
                data = governance_service.export_json(session, "kb_test")
                self.assertEqual(data["total"], 2)
                for it in data["items"]:
                    self.assertNotIn("path", it)
                joined = "".join(str(it) for it in data["items"])
                self.assertNotIn("data_root", joined.lower())

    def test_export_xlsx_bytes_contains_headers_no_path(self):
        with _temp_db() as (engine, session):
            with tempfile.TemporaryDirectory() as data_root:
                _seed(session, data_root)
                content = governance_service.export_xlsx_bytes(session, "kb_test")
                self.assertTrue(content.startswith(b"PK"), "xlsx 是 zip 容器")
                from io import BytesIO
                from openpyxl import load_workbook
                wb = load_workbook(BytesIO(content))
                ws = wb.active
                headers = [c.value for c in ws[1]]
                self.assertIn("filename", headers)
                self.assertNotIn("path", headers)
                filenames = [row[1].value for row in ws.iter_rows(min_row=2)]
                self.assertIn("a.pdf", filenames)

    def test_safe_id_rejects_traversal(self):
        for bad in ("..", "../x", "a/b", "a\\b", ".", "", None):
            with self.assertRaises(GovernanceError, msg=f"id={bad!r}"):
                governance_service._safe_id(bad)
        self.assertEqual(governance_service._safe_id("kb_abc123"), "kb_abc123")


class GovernanceVersionTests(unittest.TestCase):
    def test_snapshot_creates_version_with_hash_and_size(self):
        with _temp_db() as (engine, session):
            with tempfile.TemporaryDirectory() as data_root:
                _set_roots(data_root)
                try:
                    _seed(session, data_root)
                    v = governance_service.create_snapshot(
                        session, "kb_test", "file_a", creator="admin", note="首次快照"
                    )
                    self.assertEqual(v["version"], 1)
                    self.assertEqual(v["note"], "首次快照")
                    self.assertEqual(v["created_by"], "admin")
                    self.assertEqual(v["file_size"], len("content a.pdf"))
                    self.assertEqual(len(v["sha256"]), 64)
                    # 源文件副本以内容寻址方式位于受控目录，且校验一致
                    p = governance_service._blob_path(v["sha256"])
                    self.assertTrue(os.path.isfile(p))
                    self.assertEqual(governance_service._sha256_file(p), v["sha256"])
                finally:
                    _restore_roots()

    def test_versions_monotonic_no_duplicate(self):
        with _temp_db() as (engine, session):
            with tempfile.TemporaryDirectory() as data_root:
                _set_roots(data_root)
                try:
                    _seed(session, data_root)
                    v1 = governance_service.create_snapshot(session, "kb_test", "file_a")
                    # 改变内容后再次快照
                    p = os.path.join(data_root, "a.pdf")
                    Path(p).write_bytes(b"changed content")
                    v2 = governance_service.create_snapshot(session, "kb_test", "file_a")
                    v3 = governance_service.create_snapshot(session, "kb_test", "file_a")
                    self.assertEqual([v1["version"], v2["version"], v3["version"]], [1, 2, 3])
                    # 直接预插 version=1 已提交后再建快照 → 不重号
                    session.add(
                        KnowledgeDocumentVersion(
                            db_id="kb_test", file_id="file_b", version=1,
                            sha256="x", file_size=0, note="pre",
                        )
                    )
                    session.commit()
                    vb = governance_service.create_snapshot(session, "kb_test", "file_b")
                    self.assertEqual(vb["version"], 2)
                finally:
                    _restore_roots()

    def test_same_content_snapshot_dedups_copy(self):
        with _temp_db() as (engine, session):
            with tempfile.TemporaryDirectory() as data_root:
                _set_roots(data_root)
                try:
                    _seed(session, data_root)
                    v1 = governance_service.create_snapshot(session, "kb_test", "file_a")
                    v2 = governance_service.create_snapshot(session, "kb_test", "file_a")
                    self.assertTrue(v2["deduplicated"], "相同内容不重复复制")
                    self.assertEqual(v2["sha256"], v1["sha256"])
                    # 内容寻址：两个版本共享同一个 blob，磁盘只有一份，且校验一致
                    blobs = _blob_files(data_root)
                    self.assertEqual(len(blobs), 1, "相同内容只应存一份 blob")
                    self.assertEqual(blobs[0], governance_service._blob_path(v1["sha256"]))
                    self.assertEqual(
                        governance_service._sha256_file(blobs[0]), v1["sha256"]
                    )
                finally:
                    _restore_roots()

    def test_list_versions_orders_and_counts(self):
        with _temp_db() as (engine, session):
            with tempfile.TemporaryDirectory() as data_root:
                _set_roots(data_root)
                try:
                    _seed(session, data_root)
                    governance_service.create_snapshot(session, "kb_test", "file_a", note="n1")
                    governance_service.create_snapshot(session, "kb_test", "file_a", note="n2")
                    data = governance_service.list_versions(session, "kb_test", "file_a")
                    self.assertEqual(data["total"], 2)
                    self.assertEqual([v["version"] for v in data["items"]], [2, 1])
                finally:
                    _restore_roots()

    def test_version_download_dedup_chain_and_permissions(self):
        with _temp_db() as (engine, session):
            with tempfile.TemporaryDirectory() as data_root:
                _set_roots(data_root)
                try:
                    _seed(session, data_root)
                    governance_service.create_snapshot(session, "kb_test", "file_a")
                    v2 = governance_service.create_snapshot(session, "kb_test", "file_a")  # dedup
                    info = governance_service.resolve_version_download(
                        session, "kb_test", "file_a", v2["version"], _User("user")
                    )
                    self.assertEqual(info["filename"], "a.pdf")
                    self.assertGreater(info["size_bytes"], 0)
                    # restricted → 非 superadmin 拒绝
                    governance_service.update_governance(
                        session, "kb_test", "file_a", {"confidentiality": "restricted"}
                    )
                    with self.assertRaises(GovernanceForbidden):
                        governance_service.resolve_version_download(
                            session, "kb_test", "file_a", v2["version"], _User("admin")
                        )
                    info = governance_service.resolve_version_download(
                        session, "kb_test", "file_a", v2["version"], _User("superadmin")
                    )
                    self.assertEqual(info["version"], v2["version"])
                    # download_allowed=0 → 一律拒绝
                    governance_service.update_governance(
                        session, "kb_test", "file_a",
                        {"confidentiality": "internal", "download_allowed": False},
                    )
                    with self.assertRaises(GovernanceForbidden):
                        governance_service.resolve_version_download(
                            session, "kb_test", "file_a", v2["version"], _User("superadmin")
                        )
                finally:
                    _restore_roots()

    def test_version_not_found_cases(self):
        with _temp_db() as (engine, session):
            with tempfile.TemporaryDirectory() as data_root:
                _set_roots(data_root)
                try:
                    _seed(session, data_root)
                    with self.assertRaises(GovernanceNotFound):
                        governance_service.list_versions(session, "kb_test", "file_missing")
                    governance_service.create_snapshot(session, "kb_test", "file_a")
                    with self.assertRaises(GovernanceNotFound):
                        governance_service.resolve_version_download(
                            session, "kb_test", "file_a", 99, _User("superadmin")
                        )
                finally:
                    _restore_roots()

    def test_snapshot_rejects_path_traversal_source(self):
        with _temp_db() as (engine, session):
            with tempfile.TemporaryDirectory() as data_root:
                _set_roots(data_root)
                try:
                    evil = os.path.join(data_root, "..", "outside.txt")
                    Path(evil).write_bytes(b"secret")
                    _seed_file(session, data_root, "kb_test", "file_evil", "evil.txt", evil)
                    with self.assertRaises(GovernanceNotFound):
                        governance_service.create_snapshot(session, "kb_test", "file_evil")
                finally:
                    _restore_roots()

    def test_snapshot_metadata_has_no_path(self):
        with _temp_db() as (engine, session):
            with tempfile.TemporaryDirectory() as data_root:
                _set_roots(data_root)
                try:
                    _seed(session, data_root)
                    v = governance_service.create_snapshot(session, "kb_test", "file_a")
                    self.assertNotIn("path", v)
                    self.assertNotIn("path", v["sha256"])
                    row = (
                        session.query(KnowledgeDocumentVersion)
                        .filter_by(db_id="kb_test", file_id="file_a", version=v["version"])
                        .first()
                    )
                    self.assertNotIn("path", row.metadata_snapshot)
                finally:
                    _restore_roots()

    def test_snapshot_copy_failure_leaves_no_version_record(self):
        """P1-6 验收：复制失败（权限/磁盘满）→ 抛错、版本列表不增加、不留半成品目录。"""
        with _temp_db() as (engine, session):
            with tempfile.TemporaryDirectory() as data_root:
                _set_roots(data_root)
                try:
                    _seed(session, data_root)
                    with mock.patch.object(
                        governance_service.shutil, "copy2", side_effect=OSError("权限不足")
                    ):
                        with self.assertRaises(GovernanceError):
                            governance_service.create_snapshot(session, "kb_test", "file_a")
                    data = governance_service.list_versions(session, "kb_test", "file_a")
                    self.assertEqual(data["total"], 0)
                    # 磁盘不留孤儿：无已发布 blob、无残留暂存目录
                    self.assertEqual(_blob_files(data_root), [])
                finally:
                    _restore_roots()

    def test_snapshot_rename_failure_leaves_no_version_record(self):
        """复制中断（原子改名失败）同样不留可见版本记录。"""
        with _temp_db() as (engine, session):
            with tempfile.TemporaryDirectory() as data_root:
                _set_roots(data_root)
                try:
                    _seed(session, data_root)
                    with mock.patch.object(
                        governance_service.os, "replace", side_effect=OSError("写盘失败")
                    ):
                        with self.assertRaises(GovernanceError):
                            governance_service.create_snapshot(session, "kb_test", "file_a")
                    data = governance_service.list_versions(session, "kb_test", "file_a")
                    self.assertEqual(data["total"], 0)
                    # 发布失败 → 版本记录已回滚删除，磁盘不留孤儿 blob / 暂存
                    self.assertEqual(_blob_files(data_root), [])
                finally:
                    _restore_roots()

    def test_snapshot_checksum_mismatch_cleans_up_and_fails(self):
        """复制出的副本校验不一致 → 抛错、清理临时文件与版本目录。"""
        with _temp_db() as (engine, session):
            with tempfile.TemporaryDirectory() as data_root:
                _set_roots(data_root)
                try:
                    _seed(session, data_root)
                    real_sha = governance_service._sha256_file

                    def fake_sha(path):
                        if "knowledge_versions" in str(path):
                            return "0" * 64
                        return real_sha(path)

                    with mock.patch.object(governance_service, "_sha256_file", side_effect=fake_sha):
                        with self.assertRaises(GovernanceError):
                            governance_service.create_snapshot(session, "kb_test", "file_a")
                    data = governance_service.list_versions(session, "kb_test", "file_a")
                    self.assertEqual(data["total"], 0)
                    # 校验不一致 → 暂存清理，磁盘不留孤儿 blob / 暂存目录
                    self.assertEqual(_blob_files(data_root), [])
                finally:
                    _restore_roots()

    def test_dedup_corrupt_blob_rejected_no_new_record(self):
        """P1-6 验收：去重引用的既有 blob 已损坏 → 接口失败且版本列表不增加。"""
        with _temp_db() as (engine, session):
            with tempfile.TemporaryDirectory() as data_root:
                _set_roots(data_root)
                try:
                    _seed(session, data_root)
                    v1 = governance_service.create_snapshot(session, "kb_test", "file_a")
                    self.assertEqual(v1["version"], 1)
                    # 破坏既有内容寻址 blob 内容
                    blob = governance_service._blob_path(v1["sha256"])
                    self.assertTrue(os.path.isfile(blob))
                    Path(blob).write_bytes(b"tampered content")
                    with self.assertRaises(GovernanceError) as ctx:
                        governance_service.create_snapshot(session, "kb_test", "file_a")
                    self.assertIn("损坏", str(ctx.exception))
                    # 版本列表不增加，且不指向损坏 blob
                    data = governance_service.list_versions(session, "kb_test", "file_a")
                    self.assertEqual(data["total"], 1)
                    self.assertEqual(data["items"][0]["version"], 1)
                finally:
                    _restore_roots()

    def test_concurrent_two_session_create_no_lost_versions(self):
        """P1-3：两会话并发创建同一文档快照 → 版本不丢失、不重号、无断链/互删。

        用 barrier 让两个并发请求在计算版本号时同时出发，确定性触发
        (db_id, file_id, version) 唯一约束冲突 → IntegrityError 重试路径。
        断言：两个版本都落库且互不相同、都可下载、blob 校验正确、磁盘无孤儿。
        """
        with tempfile.TemporaryDirectory() as data_root:
            _set_roots(data_root)
            engine = None
            try:
                db_path = Path(data_root) / "concurrent.db"
                engine = create_engine(
                    f"sqlite:///{db_path}",
                    connect_args={"timeout": 10, "check_same_thread": False},
                )
                # WAL：并发读者/写者不互相阻塞，保证能确定性走到唯一约束冲突
                with engine.connect() as conn:
                    conn.execute(text("PRAGMA journal_mode=WAL"))
                Base.metadata.create_all(engine)
                s1 = sessionmaker(bind=engine)()
                s2 = sessionmaker(bind=engine)()
                _seed(s1, data_root)  # 提交后 s2 可见
                # 预建治理行，避免两个并发请求同时创建 (db_id,file_id) 唯一行
                s1.add(KnowledgeGovernance(db_id="kb_test", file_id="file_a"))
                s1.commit()

                real_next = governance_service._next_version_number
                gate = threading.Barrier(3)
                calls = 0

                def synced_next(session, db_id, file_id):
                    nonlocal calls
                    if calls < 2:  # 只同步两个线程各自的第一次版本号计算
                        gate.wait(timeout=10)
                    calls += 1
                    return real_next(session, db_id, file_id)

                results = []
                errors = []

                def worker(session):
                    try:
                        with mock.patch.object(
                            governance_service, "_next_version_number", side_effect=synced_next
                        ):
                            results.append(
                                governance_service.create_snapshot(session, "kb_test", "file_a")
                            )
                    except Exception as exc:  # noqa: BLE001
                        errors.append(exc)
                    finally:
                        session.close()

                t1 = threading.Thread(target=worker, args=(s1,))
                t2 = threading.Thread(target=worker, args=(s2,))
                t1.start()
                t2.start()
                gate.wait(timeout=10)
                t1.join(timeout=15)
                t2.join(timeout=15)

                self.assertEqual(errors, [], "并发创建不应失败: {}".format(errors))
                self.assertEqual(len(results), 2)
                versions = sorted(v["version"] for v in results)
                self.assertEqual(versions, [1, 2], "并发创建必须得到两个不同版本号")

                # 版本表与磁盘无断链、无孤儿：每个版本都可解析出校验一致的 blob
                check = sessionmaker(bind=engine)()
                rows = (
                    check.query(KnowledgeDocumentVersion)
                    .filter_by(db_id="kb_test", file_id="file_a")
                    .all()
                )
                self.assertEqual(len(rows), 2)
                for r in rows:
                    p = governance_service._resolve_version_blob(check, "kb_test", "file_a", r)
                    self.assertIsNotNone(p, "版本 {} 存在断链".format(r.version))
                    self.assertEqual(governance_service._sha256_file(p), r.sha256)
                check.close()
                # 同内容 → 磁盘只有一份共享 blob
                self.assertEqual(len(_blob_files(data_root)), 1, "并发同内容应共享一个 blob")
            finally:
                if engine is not None:
                    engine.dispose()  # 释放并发测试的 SQLite 文件句柄，避免 Windows 清理失败
                _restore_roots()

    def test_commit_generic_sqla_error_rolls_back_cleanly(self):
        """P1-3 + H2.2：普通 SQLAlchemyError 提交失败 → 回滚、不留版本记录。

        提交在 blob 发布之后才发生（H2 blob-first），因此提交失败会留下一个
        无引用的孤儿 blob——它可被 gc_unreferenced_blobs 清理，但绝不能留下
        浏览器可见的版本记录（记录存在而文件不存在 = 断链）。
        """
        with _temp_db() as (engine, session):
            with tempfile.TemporaryDirectory() as data_root:
                _set_roots(data_root)
                try:
                    _seed(session, data_root)
                    with mock.patch.object(
                        session, "commit",
                        side_effect=_SQAError("simulated generic commit failure"),
                    ):
                        with self.assertRaises(_SQAError):
                            governance_service.create_snapshot(session, "kb_test", "file_a")
                    # 无版本记录（无断链），但磁盘留有可被 GC 清理的孤儿 blob
                    data = governance_service.list_versions(session, "kb_test", "file_a")
                    self.assertEqual(data["total"], 0, "提交失败不得留下版本记录")
                    orphan_blobs = _blob_files(data_root)
                    self.assertEqual(len(orphan_blobs), 1, "提交失败允许留下 GC 可清理的孤儿 blob")
                    # 会话未损坏，可继续正常创建 → blob 被版本记录引用
                    governance_service.create_snapshot(session, "kb_test", "file_a")
                    self.assertEqual(
                        governance_service.list_versions(session, "kb_test", "file_a")["total"], 1
                    )
                    # 被引用的 blob 不能再被 GC 删除
                    stats = governance_service.gc_unreferenced_blobs(session, min_age_seconds=0)
                    self.assertEqual(stats["removed"], 0, "被版本引用的 blob 不得被 GC 删除")
                    self.assertEqual(len(_blob_files(data_root)), 1)
                finally:
                    _restore_roots()


def _write_orphan_blob(data_root, content, age_seconds):
    """直接向内容寻址路径写入一个无版本引用的孤儿 blob（模拟发布成功但提交失败）。"""
    sha = hashlib.sha256(content).hexdigest()
    blob = governance_service._blob_path(sha)
    os.makedirs(os.path.dirname(blob), exist_ok=True)
    Path(blob).write_bytes(content)
    old = time.time() - age_seconds
    os.utime(blob, (old, old))
    return sha


class GovernanceVersionH2AtomicityTests(unittest.TestCase):
    """H2：blob-first 原子发布 + 故障注入 + 安全 GC。

    覆盖文档 §8 H2：
    - H2.1/H2.3 发布失败 → 数据库无版本记录，不依赖第二次 DB 提交去删除；
    - H2.4 发布失败且（模拟）清理/删除提交也失败 → 最终版本数仍为 0、无断链记录；
    - H2.2 提交失败 → 允许留下 GC 可清理的孤儿 blob，但绝无版本记录存在而文件不存在；
    - H2.5 GC 只删「超过安全时间且无版本引用」的 blob，保护并发创建中的新 blob，
      且绝不触碰 _staging。
    """

    def test_publish_failure_no_version_record_no_cleanup_commit(self):
        """H2.1/H2.3/H2.4：blob 发布失败 → 无版本记录；即便模拟清理/删除提交
        也失败（session.delete 一调就抛），也不会有第二次 DB 提交可失败——
        因为发布失败时版本记录根本尚未写入。"""
        with _temp_db() as (engine, session):
            with tempfile.TemporaryDirectory() as data_root:
                _set_roots(data_root)
                try:
                    _seed(session, data_root)
                    with (
                        mock.patch.object(
                            governance_service,
                            "_publish_version_blob",
                            side_effect=OSError("模拟发布失败（磁盘满）"),
                        ),
                        # 旧实现发布失败后靠 session.delete 做清理提交；新实现绝不该
                        # 走到这一步。一旦走到，立即抛 AssertionError 使测试失败。
                        mock.patch.object(
                            session,
                            "delete",
                            side_effect=AssertionError("不应调用 session.delete 做清理"),
                        ),
                    ):
                        with self.assertRaises(GovernanceError) as ctx:
                            governance_service.create_snapshot(session, "kb_test", "file_a")
                    self.assertIn("发布失败", str(ctx.exception))
                    # 版本表为空，无断链记录
                    self.assertEqual(
                        session.query(KnowledgeDocumentVersion)
                        .filter_by(db_id="kb_test", file_id="file_a")
                        .count(),
                        0,
                    )
                    data = governance_service.list_versions(session, "kb_test", "file_a")
                    self.assertEqual(data["total"], 0)
                    # 磁盘无已发布 blob、无暂存残留
                    self.assertEqual(_blob_files(data_root), [])
                finally:
                    _restore_roots()

    def test_commit_failure_leaves_orphan_blob_gc_recovers(self):
        """H2.2：blob 已发布但版本记录提交失败 → 无版本记录、有孤儿 blob；
        GC 可回收孤儿，被引用的 blob 不受影响。"""
        with _temp_db() as (engine, session):
            with tempfile.TemporaryDirectory() as data_root:
                _set_roots(data_root)
                try:
                    _seed(session, data_root)
                    with mock.patch.object(
                        session, "commit", side_effect=_SQAError("模拟提交失败")
                    ):
                        with self.assertRaises(_SQAError):
                            governance_service.create_snapshot(session, "kb_test", "file_a")
                    self.assertEqual(
                        session.query(KnowledgeDocumentVersion)
                        .filter_by(db_id="kb_test", file_id="file_a")
                        .count(),
                        0,
                        "提交失败不得留下版本记录",
                    )
                    orphan = _blob_files(data_root)
                    self.assertEqual(len(orphan), 1, "提交失败应留下 1 个可 GC 的孤儿 blob")
                    # 孤儿 blob 已超过安全时间 → GC 可回收
                    os.utime(orphan[0], (time.time() - 7200, time.time() - 7200))
                    stats = governance_service.gc_unreferenced_blobs(
                        session, min_age_seconds=3600
                    )
                    self.assertEqual(stats["removed"], 1)
                    self.assertEqual(_blob_files(data_root), [])
                finally:
                    _restore_roots()

    def test_gc_removes_only_old_unreferenced_blobs_protects_recent(self):
        """H2.5：GC 只删超过安全时间且无版本引用的 blob；被版本引用的与
        保护期（并发创建中）的孤儿都保留。"""
        with _temp_db() as (engine, session):
            with tempfile.TemporaryDirectory() as data_root:
                _set_roots(data_root)
                try:
                    _seed(session, data_root)
                    v = governance_service.create_snapshot(session, "kb_test", "file_a")
                    referenced_sha = v["sha256"]
                    old_sha = _write_orphan_blob(data_root, b"orphan old", age_seconds=7200)
                    recent_sha = _write_orphan_blob(data_root, b"orphan recent", age_seconds=60)
                    self.assertEqual(len(_blob_files(data_root)), 3)
                    stats = governance_service.gc_unreferenced_blobs(
                        session, min_age_seconds=3600
                    )
                    self.assertEqual(stats["removed"], 1, "只应删旧的孤儿 blob")
                    self.assertEqual(stats["retained"], 2, "引用 blob + 保护期孤儿保留")
                    remaining = {
                        os.path.basename(os.path.dirname(p)) for p in _blob_files(data_root)
                    }
                    self.assertIn(referenced_sha, remaining)
                    self.assertIn(recent_sha, remaining)
                    self.assertNotIn(old_sha, remaining)
                    # 放宽安全时间（0 秒）后，保护期的孤儿也可回收，被引用 blob 仍保留
                    stats = governance_service.gc_unreferenced_blobs(
                        session, min_age_seconds=0
                    )
                    self.assertEqual(stats["removed"], 1)
                    self.assertEqual(len(_blob_files(data_root)), 1)
                finally:
                    _restore_roots()

    def test_gc_never_touches_staging(self):
        """H2.5：GC 绝不触碰请求级暂存目录。"""
        with _temp_db() as (engine, session):
            with tempfile.TemporaryDirectory() as data_root:
                _set_roots(data_root)
                try:
                    _seed(session, data_root)
                    staging = governance_service._staging_root()
                    os.makedirs(staging, exist_ok=True)
                    tmp = os.path.join(staging, "file")
                    Path(tmp).write_bytes(b"in-flight staging copy")
                    os.utime(tmp, (time.time() - 7200, time.time() - 7200))
                    stats = governance_service.gc_unreferenced_blobs(
                        session, min_age_seconds=0
                    )
                    self.assertEqual(stats["removed"], 0)
                    self.assertTrue(os.path.isfile(tmp), "GC 不得删除暂存中的请求文件")
                finally:
                    _restore_roots()


class GovernanceRouterSourceTests(unittest.TestCase):
    """router 源码级验证（避免引入 src/Milvus）。"""

    def setUp(self):
        self.src = (
            Path(__file__).resolve().parents[1]
            / "server"
            / "routers"
            / "governance_router.py"
        ).read_text(encoding="utf-8")

    def test_patch_uses_exclude_unset_to_allow_clearing(self):
        """P2-2：PATCH 必须区分“未提交”与“明确清空”，允许 null 清空可选字段。"""
        self.assertIn("model_dump(exclude_unset=True)", self.src)
        self.assertNotIn("model_dump(exclude_none=True)", self.src)

    def test_uses_streaming_response_for_download(self):
        self.assertIn("StreamingResponse", self.src)
        self.assertIn("_file_chunks(info[\"abs_path\"])", self.src)

    def test_content_disposition_rfc5987(self):
        self.assertIn("filename*=UTF-8''", self.src)
        self.assertIn("quote(filename", self.src)

    def test_permission_guard_on_management_endpoints(self):
        # 管理端点：PATCH / sync / export 仅 superadmin
        self.assertGreaterEqual(self.src.count("get_superadmin_user"), 3)
        # 列表/详情/下载允许已登录用户（受控下载在服务层鉴权）
        self.assertIn("get_required_user", self.src)

    def test_no_absolute_path_in_responses(self):
        # 路由不直接读取 KnowledgeFile.path 拼接响应
        self.assertNotIn("file_row.path", self.src)
        self.assertIn("resolve_download", self.src)

    def test_audit_records_download_success_and_failure(self):
        self.assertIn("\"knowledge.download\"", self.src)
        self.assertGreaterEqual(self.src.count("\"knowledge.download\""), 2)

    def test_blob_gc_endpoint_is_superadmin_only_and_safe(self):
        """H2.5：手动 GC 端点仅 superadmin，调用安全的 gc_unreferenced_blobs，
        带 1 小时默认保护期（Query ge=0），并记录审计。"""
        self.assertIn("@router.post(\"/blobs/gc\")", self.src)
        self.assertIn("gc_unreferenced_blobs(db, min_age_seconds=min_age_seconds)", self.src)
        self.assertIn("min_age_seconds: float = Query(default=3600, ge=0)", self.src)
        self.assertIn("get_superadmin_user", self.src)
        self.assertIn("\"knowledge.version.blob_gc\"", self.src)


if __name__ == "__main__":
    unittest.main()
