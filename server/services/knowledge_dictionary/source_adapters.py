"""来源适配器（设计文档 §7）：三种来源统一为带稳定定位信息的文档节点流。

- knowledge_base_file：直接复用已解析的 KnowledgeNode，不重复解析原文件；
- knowledge_base：冻结知识库 ID、文件 ID、文件哈希与节点集合版本（快照）；
- upload：上传文件进入受控存储，白名单解析器处理，校验扩展名/签名/大小/展开大小。

节点结构（dict）：
    {"node_id", "file_ref", "text", "page_no", "sheet_name", "cell_range", "metadata"}

本模块刻意避免模块级导入 src（Milvus/模型），所有重依赖在函数内惰性导入，
保证无文件、无模型的环境下可单元测试。
"""

from __future__ import annotations

import hashlib
import io
import os
import uuid
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterator, List, Optional, Tuple

from sqlalchemy.orm import Session

from server.models.knowledge_dictionary_models import (
    KnowledgeDictionarySource,
    KnowledgeDictionaryVersion,
)
from server.models.kb_models import KnowledgeDatabase, KnowledgeFile, KnowledgeNode

from .errors import InvalidSource, NotFound, PayloadTooLarge, UnsupportedMediaType, ValidationError

# ---------------------------------------------------------------------------
# 上传文件策略（§7.3）
# ---------------------------------------------------------------------------

UPLOAD_EXTENSIONS = ("pdf", "docx", "xlsx", "csv", "txt")
# 压缩展开上限：zip 内成员解压后总字节数（防 zip bomb）
MAX_EXPANDED_BYTES = 200 * 1024 * 1024  # 200MB
MAX_NODES_PER_FILE = 5000
MAX_CHARS_PER_NODE = 6000
PARSER_VERSION = "dict-v1"

_FILE_SIGNATURES = {
    "pdf": b"%PDF",
    "docx": b"PK\x03\x04",
    "xlsx": b"PK\x03\x04",
}

_UPLOAD_ROOT = "saves/dictionary_uploads"


def upload_root_dir() -> Path:
    """上传受控存储根目录（相对 CWD 与 config.save_dir 兼容）。"""
    root = Path(os.environ.get("DICTIONARY_UPLOAD_ROOT", _UPLOAD_ROOT))
    root.mkdir(parents=True, exist_ok=True)
    return root


def validate_upload(filename: str, content: bytes) -> Tuple[str, bytes]:
    """校验扩展名、文件签名与大小；返回 (安全后缀, 内容)。"""
    if not filename or "." not in filename:
        raise UnsupportedMediaType("无法识别文件类型，请上传 PDF/DOCX/XLSX/CSV/TXT 文件")
    ext = filename.rsplit(".", 1)[1].lower()
    if ext not in UPLOAD_EXTENSIONS:
        raise UnsupportedMediaType(f"不支持的文件类型 .{ext}，支持: {', '.join(UPLOAD_EXTENSIONS)}")
    max_bytes = _max_upload_bytes()
    if len(content) > max_bytes:
        raise PayloadTooLarge(f"文件超过大小上限 {max_bytes // (1024 * 1024)}MB")
    signature = _FILE_SIGNATURES.get(ext)
    if signature is not None:
        if not content.startswith(signature):
            raise UnsupportedMediaType("文件签名与扩展名不匹配")
    if ext in ("docx", "xlsx"):
        _check_zip_expanded(content)
    return ext, content


def _max_upload_bytes() -> int:
    try:
        from server.services.upload_service import max_upload_bytes

        return int(max_upload_bytes())
    except Exception:  # pragma: no cover - 兜底
        return 100 * 1024 * 1024


def _check_zip_expanded(content: bytes) -> None:
    """校验压缩包展开大小（zip bomb 防护）。"""
    try:
        with zipfile.ZipFile(io.BytesIO(content)) as zf:
            total = sum(info.file_size for info in zf.infolist())
    except zipfile.BadZipFile as exc:
        raise UnsupportedMediaType("文件已损坏或不是有效的 Office 文档") from exc
    if total > MAX_EXPANDED_BYTES:
        raise PayloadTooLarge("文件解压后超过大小上限")


def save_upload_file(filename: str, content: bytes) -> Dict[str, Any]:
    """把上传文件写入受控存储，返回受控存储标识（不含任意绝对路径）。"""
    ext, content = validate_upload(filename, content)
    stored_id = f"{uuid.uuid4().hex}.{ext}"
    target = upload_root_dir() / stored_id
    target.write_bytes(content)
    digest = hashlib.sha256(content).hexdigest()
    return {
        "storage_ref": stored_id,
        "content_hash": digest,
        "file_name": Path(filename).name,
        "size_bytes": len(content),
        "extension": ext,
    }


def resolve_upload_path(storage_ref: str) -> Path:
    """受控解析：storage_ref 必须是裸存储 ID，拒绝任何路径成分。"""
    if not storage_ref or storage_ref != Path(storage_ref).name or Path(storage_ref).name != storage_ref:
        raise InvalidSource(f"非法存储标识: {storage_ref}")
    path = (upload_root_dir() / storage_ref).resolve()
    root = upload_root_dir().resolve()
    if not str(path).startswith(str(root)):
        raise InvalidSource(f"非法存储标识: {storage_ref}")
    if not path.is_file():
        raise NotFound("上传文件不存在或已被清理")
    return path


# ---------------------------------------------------------------------------
# 快照（§7.2：提交时冻结文件 ID 与内容哈希，发布前比对）
# ---------------------------------------------------------------------------


def _kb_file_hash(db: Session, file: KnowledgeFile) -> str:
    """文件快照哈希：基于节点文本内容（非存储的 hash 字段，文本变更必须被检测）。"""
    nodes = db.query(KnowledgeNode).filter(KnowledgeNode.file_id == file.file_id).all()
    payload = "|".join(
        sorted(f"{n.id}:{hashlib.sha256((n.text or '').encode('utf-8')).hexdigest()}" for n in nodes)
    )
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def snapshot_kb_file(db: Session, db_id: str, file_id: str) -> Dict[str, Any]:
    kb = db.query(KnowledgeDatabase).filter(KnowledgeDatabase.db_id == db_id).first()
    if kb is None:
        raise NotFound(f"知识库不存在: {db_id}")
    file = (
        db.query(KnowledgeFile)
        .filter(KnowledgeFile.database_id == db_id, KnowledgeFile.file_id == file_id)
        .first()
    )
    if file is None:
        raise NotFound(f"知识库文件不存在: {file_id}")
    nodes = db.query(KnowledgeNode).filter(KnowledgeNode.file_id == file.file_id).all()
    if not nodes:
        raise InvalidSource("文件尚未解析或节点为空，无法作为字典来源")
    return {
        "source_type": "knowledge_base_file",
        "knowledge_base_id": db_id,
        "file_id": file.file_id,
        "file_name": file.filename,
        "content_hash": _kb_file_hash(db, file),
        "node_count": len(nodes),
        "snapshot_metadata": {"node_ids": sorted(n.id for n in nodes)},
    }


def snapshot_kb(db: Session, db_id: str) -> List[Dict[str, Any]]:
    kb = db.query(KnowledgeDatabase).filter(KnowledgeDatabase.db_id == db_id).first()
    if kb is None:
        raise NotFound(f"知识库不存在: {db_id}")
    files = db.query(KnowledgeFile).filter(KnowledgeFile.database_id == db_id).all()
    if not files:
        raise InvalidSource("知识库没有任何文件")
    snapshots = []
    for file in files:
        nodes = db.query(KnowledgeNode).filter(KnowledgeNode.file_id == file.file_id).all()
        if not nodes:
            continue
        snapshots.append(
            {
                "source_type": "knowledge_base",
                "knowledge_base_id": db_id,
                "file_id": file.file_id,
                "file_name": file.filename,
                "content_hash": _kb_file_hash(db, file),
                "node_count": len(nodes),
                "snapshot_metadata": {"node_ids": sorted(n.id for n in nodes)},
            }
        )
    if not snapshots:
        raise InvalidSource("知识库中没有已解析的文件")
    return snapshots


def snapshot_upload(meta: Dict[str, Any]) -> Dict[str, Any]:
    """上传来源快照：storage_ref 已由 save_upload_file 校验并落盘。"""
    path = resolve_upload_path(meta["storage_ref"])
    digest = hashlib.sha256(path.read_bytes()).hexdigest()
    return {
        "source_type": "upload",
        "storage_ref": meta["storage_ref"],
        "file_name": meta.get("file_name"),
        "content_hash": digest,
        "snapshot_metadata": {"size_bytes": path.stat().st_size},
    }


# ---------------------------------------------------------------------------
# 来源选择列表（生成向导用；admin/superadmin 可读，避免耦合 superadmin 专属的数据接口）
# ---------------------------------------------------------------------------


def list_kbs_for_source(db: Session) -> List[Dict[str, Any]]:
    """列出可用作字典来源的知识库及文件统计（§4.3）。"""
    kbs = db.query(KnowledgeDatabase).all()
    out = []
    for kb in kbs:
        files = db.query(KnowledgeFile).filter(KnowledgeFile.database_id == kb.db_id).all()
        parsed = 0
        error = 0
        for f in files:
            nodes = db.query(KnowledgeNode).filter(KnowledgeNode.file_id == f.file_id).count()
            if nodes:
                parsed += 1
            elif f.status in ("failed", "error"):
                error += 1
        out.append(
            {
                "db_id": kb.db_id,
                "name": kb.name,
                "description": kb.description,
                "file_count": len(files),
                "parsed_count": parsed,
                "error_count": error,
            }
        )
    return out


def list_kb_files_for_source(
    db: Session, db_id: str, *, keyword: str = "", page: int = 1, page_size: int = 20
) -> Dict[str, Any]:
    """知识库文件可搜索分页列表（生成向导步骤一）。"""
    kb = db.query(KnowledgeDatabase).filter(KnowledgeDatabase.db_id == db_id).first()
    if kb is None:
        raise NotFound(f"知识库不存在: {db_id}")
    q = db.query(KnowledgeFile).filter(KnowledgeFile.database_id == db_id)
    if keyword:
        q = q.filter(KnowledgeFile.filename.like(f"%{keyword}%"))
    total = q.count()
    rows = q.order_by(KnowledgeFile.id).offset((page - 1) * page_size).limit(page_size).all()
    items = []
    for f in rows:
        node_count = db.query(KnowledgeNode).filter(KnowledgeNode.file_id == f.file_id).count()
        items.append(
            {
                "file_id": f.file_id,
                "file_name": f.filename,
                "file_type": f.file_type,
                "status": f.status,
                "node_count": node_count,
            }
        )
    return {"items": items, "total": total, "page": page, "page_size": page_size}


def create_source_rows(db: Session, version: KnowledgeDictionaryVersion, snapshots: List[Dict[str, Any]]) -> str:
    """按快照创建来源行，返回整个来源集的快照哈希（发布前比对依据）。"""
    rows = []
    for snap in snapshots:
        row = KnowledgeDictionarySource(
            version_id=version.id,
            source_type=snap["source_type"],
            knowledge_base_id=snap.get("knowledge_base_id"),
            file_id=snap.get("file_id"),
            file_name=snap.get("file_name"),
            storage_ref=snap.get("storage_ref"),
            content_hash=snap.get("content_hash"),
            parser_version=PARSER_VERSION,
            snapshot_metadata=snap.get("snapshot_metadata"),
        )
        db.add(row)
        rows.append(row)
    db.flush()
    digest = _snapshot_digest(
        [
            {
                "source_type": snap["source_type"],
                "knowledge_base_id": snap.get("knowledge_base_id"),
                "file_id": snap.get("file_id"),
                "storage_ref": snap.get("storage_ref"),
                "content_hash": snap.get("content_hash"),
            }
            for snap in snapshots
        ]
    )
    version.source_snapshot_hash = digest
    return digest


def _snapshot_digest(items: List[Dict[str, Any]]) -> str:
    """来源集摘要：内部按稳定键排序，创建与发布校验两边使用同一函数保证可比较。"""
    import json

    ordered = sorted(
        items,
        key=lambda item: (
            item.get("source_type") or "",
            item.get("file_id") or item.get("storage_ref") or "",
        ),
    )
    payload = json.dumps(ordered, ensure_ascii=False, sort_keys=True)
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def _current_content_hash(db: Session, source: KnowledgeDictionarySource) -> Optional[str]:
    """计算单个来源的当前内容哈希（kb 按节点文本、upload 按文件字节）。"""
    if source.source_type in ("knowledge_base_file", "knowledge_base"):
        file = (
            db.query(KnowledgeFile)
            .filter(
                KnowledgeFile.database_id == source.knowledge_base_id,
                KnowledgeFile.file_id == source.file_id,
            )
            .first()
        )
        if file is None:
            return None
        return _kb_file_hash(db, file)
    if source.source_type == "upload":
        try:
            path = resolve_upload_path(source.storage_ref or "")
            return hashlib.sha256(path.read_bytes()).hexdigest()
        except (NotFound, InvalidSource):
            return None
    return None


def current_source_hashes(db: Session, sources: List[KnowledgeDictionarySource]) -> Dict[str, Any]:
    """发布前计算每个来源的当前哈希（§7.2：来源变化检测）。"""
    out = []
    for s in sources:
        current = _current_content_hash(db, s)
        out.append(
            {
                "source_id": s.id,
                "current_hash": current,
                "changed": current != s.content_hash,
            }
        )
    return {"sources": out, "any_changed": any(item.get("changed") for item in out)}


# ---------------------------------------------------------------------------
# 节点流（§7.1/§7.3）
# ---------------------------------------------------------------------------


def _clip_text(text: str) -> str:
    text = (text or "").strip()
    return text[:MAX_CHARS_PER_NODE]


def iter_kb_file_nodes(db: Session, source: KnowledgeDictionarySource) -> Iterator[Dict[str, Any]]:
    """知识库文件来源：按节点 ID 顺序输出已解析节点。"""
    node_ids = (source.snapshot_metadata or {}).get("node_ids") or []
    if node_ids:
        nodes = (
            db.query(KnowledgeNode)
            .filter(KnowledgeNode.id.in_(node_ids))
            .order_by(KnowledgeNode.id)
            .all()
        )
    else:
        nodes = (
            db.query(KnowledgeNode)
            .filter(KnowledgeNode.file_id == source.file_id)
            .order_by(KnowledgeNode.id)
            .all()
        )
    for node in nodes[:MAX_NODES_PER_FILE]:
        text = _clip_text(node.text)
        if not text:
            continue
        meta = node.meta_info or {}
        yield {
            "node_id": str(node.id),
            "file_ref": {"source_id": source.id, "file_name": source.file_name, "file_id": source.file_id},
            "text": text,
            "page_no": _first_str(meta.get("page_no"), meta.get("page"), meta.get("page_label")),
            "sheet_name": _first_str(meta.get("sheet_name")),
            "cell_range": _first_str(meta.get("cell_range")),
            "metadata": meta,
        }


def iter_upload_nodes(source: KnowledgeDictionarySource) -> Iterator[Dict[str, Any]]:
    """上传文件来源：按扩展名选择白名单解析器。"""
    path = resolve_upload_path(source.storage_ref or "")
    ext = path.suffix.lower().lstrip(".")
    parser = _PARSERS.get(ext)
    if parser is None:  # pragma: no cover
        raise UnsupportedMediaType(f"不支持的文件类型: .{ext}")
    for index, node in enumerate(parser(path)):
        if index >= MAX_NODES_PER_FILE:
            return
        yield {
            "node_id": f"upload:{source.id}:{index}",
            "file_ref": {"source_id": source.id, "file_name": source.file_name},
            "text": _clip_text(node["text"]),
            "page_no": node.get("page_no"),
            "sheet_name": node.get("sheet_name"),
            "cell_range": node.get("cell_range"),
            "metadata": node.get("metadata") or {},
        }


def iter_source_nodes(db: Session, source: KnowledgeDictionarySource) -> Iterator[Dict[str, Any]]:
    if source.source_type in ("knowledge_base_file", "knowledge_base"):
        yield from iter_kb_file_nodes(db, source)
    elif source.source_type == "upload":
        yield from iter_upload_nodes(source)
    else:  # pragma: no cover
        raise InvalidSource(f"未知来源类型: {source.source_type}")


def _first_str(*values: Any) -> Optional[str]:
    for value in values:
        if value is not None and str(value).strip():
            return str(value).strip()
    return None


# ---------------------------------------------------------------------------
# 白名单解析器（§7.3）
# ---------------------------------------------------------------------------


def _parse_pdf(path: Path) -> Iterator[Dict[str, Any]]:
    import fitz  # PyMuPDF

    with fitz.open(path) as doc:
        for page_index in range(len(doc)):
            text = doc[page_index].get_text("text")
            if not text.strip():
                continue
            yield {
                "text": text,
                "page_no": str(page_index + 1),
                "metadata": {"parser": "pymupdf"},
            }


def _parse_docx(path: Path) -> Iterator[Dict[str, Any]]:
    import docx2txt

    text = docx2txt.process(str(path)) or ""
    if not text.strip():
        return
    # 按段落粗分块
    for index, chunk in enumerate(_chunk_text(text, MAX_CHARS_PER_NODE)):
        yield {"text": chunk, "metadata": {"parser": "docx2txt", "chunk": index}}


def _parse_xlsx(path: Path) -> Iterator[Dict[str, Any]]:
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    for ws in wb.worksheets:
        parts: List[str] = []
        start_row = 1
        row_count = 0
        for row_index, row in enumerate(ws.iter_rows(values_only=True), start=1):
            values = ["" if v is None else str(v) for v in row]
            line = " | ".join(values).rstrip(" |")
            if not line.strip():
                if parts:
                    yield _xlsx_node(ws.title, parts, start_row, row_index - 1)
                    parts = []
                continue
            if not parts:
                start_row = row_index
            parts.append(line)
            row_count += 1
            # 每 60 行拆一个节点，避免单个单元格区域过大
            if row_count >= 60:
                yield _xlsx_node(ws.title, parts, start_row, row_index)
                parts = []
                start_row = row_index + 1
                row_count = 0
        if parts:
            yield _xlsx_node(ws.title, parts, start_row, ws.max_row or start_row)


def _xlsx_node(sheet: str, parts: List[str], start: int, end: int) -> Dict[str, Any]:
    return {
        "text": "\n".join(parts),
        "sheet_name": sheet,
        "cell_range": f"A{start}:A{end}",
        "metadata": {"parser": "openpyxl", "sheet": sheet},
    }


def _parse_csv(path: Path) -> Iterator[Dict[str, Any]]:
    import csv

    with open(path, "r", encoding="utf-8-sig", errors="replace", newline="") as fh:
        reader = csv.reader(fh)
        parts: List[str] = []
        start = 1
        for index, row in enumerate(reader, start=1):
            line = " , ".join(row).rstrip(" ,")
            if not line.strip():
                if parts:
                    yield {"text": "\n".join(parts), "cell_range": f"A{start}:A{index - 1}"}
                    parts = []
                continue
            if not parts:
                start = index
            parts.append(line)
            if len(parts) >= 60:
                yield {"text": "\n".join(parts), "cell_range": f"A{start}:A{index}"}
                parts = []
                start = index + 1
        if parts:
            yield {"text": "\n".join(parts), "cell_range": f"A{start}:A{len(parts)}"}


def _parse_txt(path: Path) -> Iterator[Dict[str, Any]]:
    text = path.read_text(encoding="utf-8", errors="replace")
    for index, chunk in enumerate(_chunk_text(text, MAX_CHARS_PER_NODE)):
        yield {"text": chunk, "metadata": {"parser": "plain", "chunk": index}}


def _chunk_text(text: str, size: int) -> List[str]:
    """按段落边界粗分块，避免截断句子。"""
    paragraphs = [p.strip() for p in text.split("\n\n") if p.strip()]
    chunks: List[str] = []
    current = ""
    for para in paragraphs:
        if len(current) + len(para) + 1 > size and current:
            chunks.append(current)
            current = para
        else:
            current = f"{current}\n\n{para}".strip()
        while len(current) > size:
            chunks.append(current[:size])
            current = current[size:]
    if current:
        chunks.append(current)
    return chunks or ([text[:size]] if text.strip() else [])


_PARSERS = {
    "pdf": _parse_pdf,
    "docx": _parse_docx,
    "xlsx": _parse_xlsx,
    "csv": _parse_csv,
    "txt": _parse_txt,
}
