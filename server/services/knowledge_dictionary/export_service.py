"""字典导出（设计文档 §13.5）：XLSX / CSV / JSON。

- XLSX 为默认交付格式：字典条目、来源证据、版本信息三个工作表；
- CSV 对以 = + - @ 开头的值做公式注入防护；
- 导出文件名与响应头正确处理中文。
"""

from __future__ import annotations

import csv
import io
import json
import re
from datetime import datetime, timezone
from typing import Any, Dict, List, Tuple
from urllib.parse import quote

from sqlalchemy.orm import Session

from server.models.knowledge_dictionary_models import (
    KnowledgeDictionary,
    KnowledgeDictionaryEntry,
    KnowledgeDictionaryEvidence,
    KnowledgeDictionarySource,
    KnowledgeDictionaryVersion,
)

from . import repository as repo
from .errors import NotFound, ValidationError
from .permissions import ensure_manager

_FORMAT = ("xlsx", "csv", "json")

# 公式注入前缀（OWASP）
_FORMULA_PREFIX_RE = re.compile(r"^[=+\-@\t\r]")


def _csv_safe(value: Any) -> str:
    """CSV 公式注入防护：以 = + - @ 开头的值加单引号前缀。"""
    text = "" if value is None else str(value)
    if _FORMULA_PREFIX_RE.match(text):
        return "'" + text
    return text


def _entries_as_dicts(entries: List[KnowledgeDictionaryEntry]) -> List[Dict[str, Any]]:
    rows = []
    for e in entries:
        rows.append(
            {
                "id": e.id,
                "category": e.category or "",
                "standard_name": e.standard_name,
                "definition": e.definition,
                "unit": e.unit or "",
                "data_type": e.data_type,
                "synonyms": ", ".join(json.loads(e.synonyms) if isinstance(e.synonyms, str) else (e.synonyms or [])),
                "value_rule": e.value_rule or "",
                "review_status": e.review_status,
                "confidence": round(float(e.confidence or 0), 4),
            }
        )
    return rows


def _evidence_rows(db: Session, version_id: int) -> List[Dict[str, Any]]:
    rows = (
        db.query(KnowledgeDictionaryEvidence, KnowledgeDictionarySource)
        .join(
            KnowledgeDictionarySource,
            KnowledgeDictionaryEvidence.source_id == KnowledgeDictionarySource.id,
            isouter=True,
        )
        .join(
            KnowledgeDictionaryEntry,
            KnowledgeDictionaryEvidence.entry_id == KnowledgeDictionaryEntry.id,
        )
        .filter(KnowledgeDictionaryEntry.version_id == version_id)
        .order_by(KnowledgeDictionaryEvidence.entry_id, KnowledgeDictionaryEvidence.id)
        .all()
    )
    out = []
    for ev, source in rows:
        out.append(
            {
                "entry_id": ev.entry_id,
                "field_path": ev.field_path or "",
                "quote": ev.quote,
                "page_no": ev.page_no or "",
                "sheet_name": ev.sheet_name or "",
                "cell_range": ev.cell_range or "",
                "inferred": "是" if ev.inferred else "否",
                "source_file": source.file_name if source is not None else "",
            }
        )
    return out


def export_version(
    db: Session,
    user: Any,
    dictionary_id: int,
    version_id: int,
    *,
    fmt: str = "xlsx",
    include_rejected: bool = False,
) -> Tuple[bytes, str, str]:
    """导出指定版本。返回 (content, media_type, download_filename)。"""
    ensure_manager(user)
    if fmt not in _FORMAT:
        raise ValidationError(f"不支持的导出格式: {fmt}，支持 xlsx/csv/json")
    dictionary = repo.get_dictionary(db, dictionary_id)
    version = repo.get_version_of_dictionary(db, dictionary_id, version_id)
    entries = (
        db.query(KnowledgeDictionaryEntry)
        .filter(
            KnowledgeDictionaryEntry.version_id == version.id,
        )
        .order_by(KnowledgeDictionaryEntry.id)
        .all()
    )
    if not include_rejected:
        entries = [e for e in entries if e.review_status != "rejected"]

    base_name = f"{dictionary.name}-V{version.version_no}"
    safe_name = base_name.replace("/", "_").replace("\\", "_")

    if fmt == "json":
        payload = {
            "dictionary": repo.serialize_dictionary(dictionary),
            "version": repo.serialize_version(version),
            "entries": _entries_as_dicts(entries),
            "evidences": _evidence_rows(db, version.id),
            "exported_at": datetime.now(timezone.utc).isoformat(),
        }
        content = json.dumps(payload, ensure_ascii=False, indent=2).encode("utf-8")
        return content, "application/json", f"{safe_name}.json"

    if fmt == "csv":
        buffer = io.StringIO()
        fieldnames = [
            "id",
            "category",
            "standard_name",
            "definition",
            "unit",
            "data_type",
            "synonyms",
            "value_rule",
            "review_status",
            "confidence",
        ]
        writer = csv.DictWriter(buffer, fieldnames=fieldnames)
        writer.writeheader()
        for row in _entries_as_dicts(entries):
            writer.writerow({k: _csv_safe(v) for k, v in row.items()})
        content = buffer.getvalue().encode("utf-8-sig")
        return content, "text/csv", f"{safe_name}.csv"

    # xlsx：三工作表（条目 / 证据 / 版本信息）
    import openpyxl
    from openpyxl.styles import Font

    wb = openpyxl.Workbook()
    ws_entries = wb.active
    ws_entries.title = "字典条目"
    entry_headers = ["ID", "分类", "标准名称", "定义", "单位", "数据类型", "同义词", "取值规则", "审核状态", "置信度"]
    ws_entries.append(entry_headers)
    for row in _entries_as_dicts(entries):
        ws_entries.append(
            [
                row["id"],
                row["category"],
                row["standard_name"],
                row["definition"],
                row["unit"],
                row["data_type"],
                row["synonyms"],
                row["value_rule"],
                row["review_status"],
                row["confidence"],
            ]
        )

    ws_evidence = wb.create_sheet("来源证据")
    evidence_headers = ["条目ID", "字段", "原文引用", "页码", "工作表", "单元格", "推断", "来源文件"]
    ws_evidence.append(evidence_headers)
    for row in _evidence_rows(db, version.id):
        ws_evidence.append(
            [
                row["entry_id"],
                row["field_path"],
                row["quote"],
                row["page_no"],
                row["sheet_name"],
                row["cell_range"],
                row["inferred"],
                row["source_file"],
            ]
        )

    ws_info = wb.create_sheet("版本信息")
    info = repo.serialize_version(version)
    info["dictionary_name"] = dictionary.name
    info["domain"] = dictionary.domain or ""
    for key in ["dictionary_name", "domain", "version_no", "status", "index_status", "entry_count", "pending_count", "conflict_count", "vector_count", "source_snapshot_hash", "embedding_config_hash", "created_at", "published_at"]:
        ws_info.append([key, info.get(key, "")])

    for ws in (ws_entries, ws_evidence, ws_info):
        for cell in ws[1]:
            cell.font = Font(bold=True)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", f"{safe_name}.xlsx"


def export_content_disposition(filename: str) -> str:
    """RFC 5987 安全编码 Content-Disposition（含中文）。"""
    fallback = "".join(c for c in filename if ord(c) < 128) or "dictionary"
    encoded = quote(filename, safe="")
    return f"attachment; filename=\"{fallback}\"; filename*=UTF-8''{encoded}"
