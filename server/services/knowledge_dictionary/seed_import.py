"""XinJiang 种子迁移（设计文档 §9）：旧版压裂字典资产一次性幂等迁移为种子字典。

- 读取 XinJiang/extract/水平井压裂数据管理.xlsx 的多行表头（分组/字段名/单位/示例值）；
- 合并旧脚本中的同义词、值字典和线索字典（以数据常量内嵌，旧脚本不作为生产运行时依赖）；
- 迁移结果保存为「压裂知识字典 V1」草稿版本，经管理员审核、索引、发布后使用；
- 幂等：保存导入器版本和源文件哈希，相同版本/哈希重复执行不创建重复字典或条目；
  源文件变化时创建新草稿版本。
"""

from __future__ import annotations

import hashlib
import re
from functools import lru_cache
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional, Set

from sqlalchemy.orm import Session

from server.models.knowledge_dictionary_models import (
    KnowledgeDictionary,
    KnowledgeDictionaryEntry,
    KnowledgeDictionaryEvidence,
    KnowledgeDictionarySource,
    KnowledgeDictionaryVersion,
)

from . import repository as repo
from .errors import Conflict, NotFound, ValidationError
from .normalizer import compute_confidence, content_hash, normalize_name, normalize_synonyms, normalize_unit
from .permissions import ensure_manager

SEED_DICTIONARY_NAME = "压裂知识字典 V1"
SEED_IMPORTER_VERSION = "xinjiang-v1"
SEED_DOMAIN = "石油工程-压裂"

_PROJECT_ROOT = Path(__file__).resolve().parents[3]
SEED_XLSX_PATH = _PROJECT_ROOT / "XinJiang" / "extract" / "水平井压裂数据管理.xlsx"

# 工作表顺序与旧脚本 filter_util.sheet_index 对齐
_SHEETS = ("基础表", "设计表", "施工表", "生产表")

# ---------------------------------------------------------------------------
# 旧脚本数据内嵌（filter_util.py 的 synonym_dict / filter_dict / value_dict / clue_dict）
# ---------------------------------------------------------------------------

_SEED_SYNONYMS: Dict[int, Dict[str, List[str]]] = {
    0: {
        "大地坐标X": ["井位坐标，纵(X)", "纵坐标(X)"],
        "大地坐标Y": ["井位坐标，横(Y)", "横坐标（Y）"],
        "A测深": ["A点测深", "A靶点测深", "A点斜深", "A靶点斜深", "A点", "A靶点"],
        "B测深": ["B点测深", "B靶点测深", "B点斜深", "B靶点斜深", "B点", "B靶点"],
        "A垂深": ["A点垂深", "A靶点垂深"],
        "B垂深": ["B点垂深", "B靶点垂深"],
        "水平段长": ["水平段长度"],
        "井口装置": ["完井井口"],
        "目的层系": ["完钻层位"],
        "Ⅰ类油层": ["1类油层"],
        "Ⅱ类油层": ["2类油层"],
        "Ⅲ类油层": ["3类油层"],
        "方案产能": ["产能"],
        "杨氏模量_avg": ["杨氏模量平均值"],
        "泊松比_avg": ["泊松比平均值"],
        "脆性指数": ["脆性指数平均值"],
        "孔隙度": ["解释油层孔隙度最小值"],
        "So_min": ["含油饱和度最小值", "含油饱和度"],
        "So_max": ["含油饱和度最大值"],
        "So_avg": ["含油饱和度平均值"],
        "油层温度": ["油藏中部温度", "储层温度", "温度"],
        "水平油层厚度": ["钻遇油层", "解释油层"],
        "K_min": ["渗透率最小值", "渗透率"],
        "K_max": ["渗透率最大值"],
        "K_avg": ["渗透率平均值"],
        "地层压力系数": ["压力系数无因次", "压力系数"],
        "原始地层压力": ["压力"],
        "井身结构": ["套管"],
    },
    1: {
        "前置液比": ["前置液比例"],
        "均砂比": ["平均砂比"],
        "施工正常限压": ["施工限压"],
        "费用预算": ["压裂费用预算", "合计", "压裂工程累计费用"],
        "设计总液量": ["压裂液总量", "压裂液用量总量", "总液量", "总入井液量", "设计"],
        "滑溜水、低粘比例": ["滑溜水比例", "滑溜液比例"],
        "分段工艺": ["分段"],
        "压裂工艺": ["压裂方式"],
        "均段间距": ["平均段间距", "段距平均"],
        "均簇间距": ["平均簇间距"],
        "段数": ["合计段数", "总段数", "段"],
        "簇数": ["合计簇数", "总簇数", "簇"],
        "首段射孔工艺": ["第一级射孔", "射孔"],
        "分段工具": ["桥塞"],
        "施工排量": [],
    },
}

_SEED_VALUE_EXAMPLES: Dict[int, Dict[str, List[str]]] = {
    0: {
        "目的层系": ["P2l22-3"],
        "平台/单井": ["57A"],
        "井身结构": ["二开", "三开"],
        "油层套管": ["139.7mm×12.09mm×TP110V/TP125V"],
        "井口装置": ["KY105/78-65"],
        "固井质量": ["合格", "优秀", "良好", "不合格"],
        "水敏性": ["强", "中", "弱", "无"],
        "Ⅰ类油层": ["300"],
        "Ⅱ类油层": ["400"],
        "井型": ["水平井"],
        "井别": ["采油井", "开发井"],
    },
    1: {
        "压裂工艺": ["桥塞射孔"],
        "分段工具": ["速钻桥塞", "可溶桥塞"],
        "首段射孔工艺": ["连油射孔", "连续油管传输射孔"],
        "水平井段": ["3690.00~5800.00"],
        "支撑剂类型": ["70/140+40/70+30/50目石英砂"],
        "压裂液类型": ["胍胶+滑溜水"],
        "暂堵剂类型": ["颗粒+粉末各一半"],
        "施工排量": ["15-18"],
        "石英砂": ["468"],
        "30/50目石英砂": ["1350"],
    },
}

_SEED_CLUES: Dict[int, Dict[str, str]] = {
    0: {
        "井身结构": "该字段的取值是文字，形式是一开、二开、三开，判断方法是看该井用了几个套管，如果用了表套和油套，则是二开；用了表套、技套和油套，则是三开。",
        "孔隙度": "该字段的取值是数字，指的是解释油层孔隙度的最小值，可能会直接给出，也有可能让你从范围中进行提取",
        "So_min": "该字段的取值是数字，指的是含油饱和度最小值，可能会直接给出，也可能是给出一个范围，你从范围中提取",
        "So_max": "该字段的取值是数字，指的是含油饱和度最大值，可能会直接给出，也可能是给出一个范围，你从范围中提取",
        "So_avg": "该字段的取值是数字，指的是含油饱和度的平均值，若出现在表中未明确最大最小，一般指均值",
        "水平油层厚度": "该字段的取值是数字",
        "K_min": "该字段的取值是数字，指的是渗透率最小值，可能会直接给出，也可能是给出一个范围，你从范围中提取",
        "K_max": "该字段的取值是数字，指的是渗透率最大值，可能会直接给出，也可能是给出一个范围，你从范围中提取",
        "K_avg": "该字段的取值是数字，指的是渗透率平均值",
        "目的层系": "该字段通常以（Pxlxx-x: xm-xm/xm）形式出现，其中Pxlxx-x就是目的层系",
    },
    1: {
        "施工排量": "该字段的取值是一个范围，可能需要你从表或段落中提取出最小值和最大值后以范围的形式返回",
        "石英砂": "该字段的取值是数字，指的是使用该材料的总量，而不是百分比等，通常位于类似表名为材料统计表的表中或段落中",
        "陶粒": "该字段的取值是数字，指的是使用该材料的总量，而不是百分比等，通常位于类似表名为材料统计表的表中或段落中",
        "费用预算": "该字段通常位于表名位于费用预算表的表中或位于段落中，指的是合计费用",
        "支撑剂类型": "该字段通常位于表名类似支撑剂选择表的表中，或者位于有关支撑剂选择的段落中，指的是使用了哪些材料作为支撑剂",
        "分段工具": "该字段指的是使用的分段工具的类型，要分清使用的桥塞到底是可溶桥塞还是速钻桥塞",
    },
}


# ---------------------------------------------------------------------------
# xlsx 解析（多行表头：第 1 行分组、第 2 行字段名、第 3 行单位）
# ---------------------------------------------------------------------------

_UNIT_WRAP_RE = re.compile(r"^[（(](.*)[）)]$")


def _source_hash() -> str:
    if not SEED_XLSX_PATH.is_file():
        raise NotFound(f"种子文件不存在: {SEED_XLSX_PATH}")
    return hashlib.sha256(SEED_XLSX_PATH.read_bytes()).hexdigest()


def _column_letter(index: int) -> str:
    """0 基列号转 Excel 列字母。"""
    letters = ""
    index += 1
    while index:
        index, rem = divmod(index - 1, 26)
        letters = chr(65 + rem) + letters
    return letters


def parse_seed_xlsx() -> List[Dict[str, Any]]:
    """解析种子 xlsx：每列产出 (sheet, group, name, unit, value_example)。

    分组（第 1 行）向前填充；同名列（多级列头）按分组+名+单位去重；
    名称为空或纯空白列跳过。
    """
    import openpyxl

    if not SEED_XLSX_PATH.is_file():
        raise NotFound(f"种子文件不存在: {SEED_XLSX_PATH}")
    wb = openpyxl.load_workbook(SEED_XLSX_PATH, data_only=True, read_only=True)
    entries: List[Dict[str, Any]] = []
    seen_keys = set()
    for sheet_index, sheet_name in enumerate(_SHEETS):
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(min_row=1, max_row=4, values_only=True))
        if len(rows) < 2:
            continue
        group_row, name_row, unit_row = rows[0], rows[1], rows[2] if len(rows) > 2 else None
        data_row = rows[3] if len(rows) > 3 else None
        current_group = ""
        for col_index in range(len(name_row)):
            group_value = group_row[col_index] if col_index < len(group_row) else None
            if group_value is not None and str(group_value).strip():
                current_group = str(group_value).strip()
            name = name_row[col_index]
            if name is None or not str(name).strip():
                continue
            name = str(name).strip()
            unit = None
            if unit_row is not None and col_index < len(unit_row) and unit_row[col_index] is not None:
                unit = str(unit_row[col_index]).strip() or None
            if unit:
                match = _UNIT_WRAP_RE.match(unit)
                if match:
                    unit = match.group(1)
            example = None
            if data_row is not None and col_index < len(data_row) and data_row[col_index] is not None:
                example = str(data_row[col_index]).strip() or None
            # 同表同分组同名同单位只保留首列（多级列头去重）
            key = (sheet_index, current_group, name, unit or "")
            if key in seen_keys:
                continue
            seen_keys.add(key)
            entries.append(
                {
                    "sheet_index": sheet_index,
                    "sheet_name": sheet_name,
                    "category": current_group or sheet_name,
                    "name": name,
                    "unit": unit,
                    "column": _column_letter(col_index),
                    "example": example,
                }
            )
    return entries


# ---------------------------------------------------------------------------
# 种子条目构建
# ---------------------------------------------------------------------------


def build_seed_entries() -> List[Dict[str, Any]]:
    """把 xlsx 列与旧脚本字典合并为条目草稿（不含数据库写入）。"""
    columns = parse_seed_xlsx()
    out: List[Dict[str, Any]] = []
    for col in columns:
        sheet_index = col["sheet_index"]
        name = col["name"]
        unit = col["unit"]
        synonyms: List[str] = []
        if unit:
            synonyms.append(f"{name}({unit})")
        # 旧脚本同义词：按 名 / 名(单位) / 分组别名 匹配
        for key, values in _SEED_SYNONYMS.get(sheet_index, {}).items():
            if key in (name, f"{name}({unit})", f"{name}{unit}") or name.startswith(key) or key.startswith(name):
                synonyms.extend(values)
        clue = None
        for key, text in _SEED_CLUES.get(sheet_index, {}).items():
            if key in (name, f"{name}({unit})") or (key and (name.startswith(key) or key.startswith(name))):
                clue = text
                break
        examples: List[str] = []
        for key, values in _SEED_VALUE_EXAMPLES.get(sheet_index, {}).items():
            if key in (name, f"{name}({unit})") or (key and (name.startswith(key) or key.startswith(name))):
                examples.extend(values)
        definition = clue or f"水平井压裂数据管理「{col['sheet_name']}」表中的字段「{name}」" + (
            f"，单位 {unit}" if unit else ""
        )
        value_rule = None
        if examples:
            value_rule = "取值示例: " + "、".join(str(v) for v in examples[:8])
        elif col.get("example"):
            value_rule = f"取值示例: {col['example']}"
        entry = {
            "category": col["category"],
            "standard_name": name,
            "definition": definition,
            "unit": unit,
            "data_type": "string",
            "synonyms": synonyms,
            "value_rule": value_rule,
            "review_status": "pending",
            "confidence": compute_confidence(
                {
                    "has_definition": bool(clue),
                    "explicit_unit": bool(unit),
                    "explicit_type": False,
                    "seed_hit": True,
                    "multi_source": False,
                    "complete": bool(unit and clue),
                }
            ),
            "evidence": [
                {
                    "node_id": f"seed:{col['sheet_name']}:{col['column']}",
                    "quote": name,
                    "field_path": "standard_name",
                    "sheet_name": col["sheet_name"],
                    "cell_range": f"{col['column']}2:{col['column']}2",
                    "inferred": False,
                }
            ],
        }
        out.append(entry)
    return out


@lru_cache(maxsize=2)
def load_seed_names() -> Set[str]:
    """全部种子标准名与同义词的规范化集合（生成管线 seed_hit 检测用）。"""
    names: Set[str] = set()
    for entry in build_seed_entries():
        names.add(normalize_name(entry["standard_name"]))
        for syn in entry["synonyms"]:
            names.add(normalize_name(syn))
    return {n for n in names if n}


# ---------------------------------------------------------------------------
# 幂等导入（§9）
# ---------------------------------------------------------------------------


def _find_seed_dictionary(db: Session) -> Optional[KnowledgeDictionary]:
    return (
        db.query(KnowledgeDictionary)
        .filter(KnowledgeDictionary.name == SEED_DICTIONARY_NAME, KnowledgeDictionary.is_deleted == 0)
        .first()
    )


def _same_seed_meta(meta: Any, version: str, digest: str) -> bool:
    if not isinstance(meta, dict):
        return False
    return meta.get("importer_version") == version and meta.get("source_hash") == digest


def import_seed_sync(
    db: Session,
    user: Any,
    *,
    force_new_version: bool = False,
    heartbeat: Optional[Callable[..., None]] = None,
) -> Dict[str, Any]:
    """幂等种子导入：相同版本+哈希返回既有字典；源文件变化时创建新草稿版本。"""
    ensure_manager(user)
    version = SEED_IMPORTER_VERSION
    digest = _source_hash()
    existing = _find_seed_dictionary(db)
    if existing is not None and not force_new_version:
        if _same_seed_meta(existing.seed_meta, version, digest):
            return {"dictionary_id": existing.id, "created": False, "reason": "种子未变化，已存在"}
        if _has_mutable_version(db, existing.id):
            raise Conflict("种子字典已存在未完成的草稿版本，请先完成或删除后再迁移")

    dictionary = existing
    if dictionary is None:
        dictionary = KnowledgeDictionary(
            name=SEED_DICTIONARY_NAME,
            description="由 XinJiang 旧版水平井压裂数据管理表迁移生成的首个压裂知识字典种子",
            domain=SEED_DOMAIN,
            status="draft",
            created_by=user.id,
            updated_by=user.id,
        )
        db.add(dictionary)
        db.flush()

    version_row = KnowledgeDictionaryVersion(
        dictionary_id=dictionary.id,
        version_no=repo.next_version_no(db, dictionary.id),
        status="draft",
        index_status="pending",
        created_by=user.id,
        generation_config={
            "kind": "import_seed",
            "importer_version": version,
            "source_hash": digest,
        },
    )
    db.add(version_row)
    db.flush()
    source = KnowledgeDictionarySource(
        version_id=version_row.id,
        source_type="upload",
        file_name=SEED_XLSX_PATH.name,
        storage_ref=None,
        content_hash=digest,
        parser_version=version,
        snapshot_metadata={"importer": version, "path_note": "XinJiang/extract/水平井压裂数据管理.xlsx"},
    )
    db.add(source)
    db.flush()

    total = 0
    for seed_entry in build_seed_entries():
        entry = KnowledgeDictionaryEntry(
            version_id=version_row.id,
            category=seed_entry["category"],
            standard_name=seed_entry["standard_name"],
            normalized_name=normalize_name(seed_entry["standard_name"]),
            definition=seed_entry["definition"],
            unit=seed_entry["unit"],
            normalized_unit=normalize_unit(seed_entry["unit"]),
            data_type=seed_entry["data_type"],
            synonyms=normalize_synonyms(seed_entry["synonyms"]),
            value_rule=seed_entry["value_rule"],
            review_status="pending",
            confidence=float(seed_entry["confidence"]),
            index_status="pending",
            created_by=user.id,
        )
        entry.content_hash = content_hash(
            {
                "category": entry.category,
                "standard_name": entry.standard_name,
                "definition": entry.definition,
                "unit": entry.unit,
                "data_type": entry.data_type,
                "synonyms": entry.synonyms,
                "value_rule": entry.value_rule,
            }
        )
        db.add(entry)
        db.flush()
        for ev in seed_entry["evidence"]:
            db.add(
                KnowledgeDictionaryEvidence(
                    entry_id=entry.id,
                    source_id=source.id,
                    node_id=ev["node_id"],
                    field_path=ev["field_path"],
                    quote=ev["quote"],
                    sheet_name=ev.get("sheet_name"),
                    cell_range=ev.get("cell_range"),
                    inferred=0,
                    evidence_hash=hashlib.sha256(ev["quote"].encode("utf-8")).hexdigest(),
                )
            )
        total += 1
        if heartbeat is not None and total % 50 == 0:
            heartbeat(stage=f"import:{total}", progress=round(min(99.0, total / 10), 2))

    repo.refresh_version_counts(db, version_row)
    dictionary.seed_meta = {"importer_version": version, "source_hash": digest, "entry_count": total}
    dictionary.updated_by = user.id
    db.commit()
    return {
        "dictionary_id": dictionary.id,
        "version_id": version_row.id,
        "version_no": version_row.version_no,
        "created": True,
        "entry_count": total,
    }


def _has_mutable_version(db: Session, dictionary_id: int) -> bool:
    return (
        db.query(KnowledgeDictionaryVersion)
        .filter(
            KnowledgeDictionaryVersion.dictionary_id == dictionary_id,
            KnowledgeDictionaryVersion.status.in_(["draft", "reviewing"]),
        )
        .count()
        > 0
    )


def create_seed_import_job(db: Session, user: Any) -> Dict[str, Any]:
    """创建 import_seed 任务（worker 执行，与 generate/index 同一套租约协议）。"""
    ensure_manager(user)
    from server.models.knowledge_dictionary_models import KnowledgeDictionaryJob

    job = KnowledgeDictionaryJob(
        job_type="import_seed",
        status="queued",
        stage="pending",
        progress=0.0,
        input_config={"importer_version": SEED_IMPORTER_VERSION},
        checkpoint={"phase": "start"},
        requested_by=user.id,
    )
    db.add(job)
    db.commit()
    db.refresh(job)
    return repo.serialize_job(job)


def import_seed_for_job(
    db: Session,
    job: Any,
    worker_id: str,
    heartbeat: Optional[Callable[..., None]] = None,
) -> None:
    """worker 侧：执行种子导入（幂等）。"""
    from server.models.user_model import User

    user = db.query(User).filter(User.id == job.requested_by).first()
    if user is None:
        raise ValidationError("任务发起人不存在")
    import_seed_sync(db, user, heartbeat=heartbeat)
